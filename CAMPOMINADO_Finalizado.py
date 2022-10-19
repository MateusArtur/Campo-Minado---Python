# Campo minado em python  #
# Nome: Laysla Rebeca Alves Ribeiro |#
# Nome: Mateus Artur                        | #

#-------------------------------------------------------------------------------------------------------------------------------------------------- #

import os 
import random 
import openpyxl

#-------------------------------------------------------------------------------------------------------------------------------------------------- #
def definirTamanhoTabuleiro():
    qtd_linhas = 0
    qtd_colunas = 0

    ''' Qtd de Linhas '''
    while qtd_linhas<3:

        qtd_linhas = int(input('Quantas linhas terá o seu tabuleiro?'))

        if qtd_linhas<3:
            print('Valor inválido. A quantidade mínima de linhas deve ser 3.')
        else:
            break
    os.system('cls')

    ''' Qtd de Colunas '''
    while qtd_colunas<3:

        qtd_colunas = int(input('Quantas colunas terá o seu tabuleiro?'))

        if qtd_colunas<3:
            print('Valor inválido. A quantidade mínima de colunas deve ser 3.')
        else:
            break
    os.system('cls')

    return qtd_linhas, qtd_colunas

#--------------------------------------------------------------------------------------------------------------------------------------------------#

def definirQtdBombas(qtd_linhas, qtd_colunas):

    ''' Definir o nível do jogo '''
    print('\n     -------------------------'  )
    print('      Escolha do Nível do Jogo   ')
    print('     __________________________ \n')
    print('| 1 - Fácil | 2 - Médio | 3 - Difícil |')
    nivel_escolhido = input('\nQual o nível de dificuldade que você deseja jogar?')
    print('\n____________________________________________________')
    print('_____________________ TABULEIRO ____________________')
    print('____________________________________________________\n')
    
    if nivel_escolhido == '1':
        qtd_bombas = 0.15 * (qtd_linhas * qtd_colunas)
        
    if nivel_escolhido == '2':
        qtd_bombas = 0.30 * (qtd_linhas * qtd_colunas)
        
    if nivel_escolhido == '3':
        qtd_bombas = 0.50 * (qtd_linhas * qtd_colunas)
    
    return int(qtd_bombas)

#--------------------------------------------------------------------------------------------------------------------------------------------------#

def criarTabuleiro(qtd_linhas, qtd_colunas):
    tabuleiro = []

    for lin in range(0, qtd_linhas):
        linha = []
        for col in range(0, qtd_colunas):
            linha.append('-')
        tabuleiro.append(linha)

    return tabuleiro
#--------------------------------------------------------------------------------------------------------------------------------------------------#

def lancarBombas(qtd_linhas, qtd_colunas, qtd_bombas, tabuleiro):
    conta_bombas = 0
 
    while (conta_bombas < qtd_bombas):
        linha_aleat = random.randint(0, qtd_linhas-1)
        coluna_aleat = random.randint(0, qtd_colunas-1)

        if tabuleiro[linha_aleat][coluna_aleat] == '-':
            tabuleiro[linha_aleat][coluna_aleat] = '*' # marcou bomba no tabuleiro
            conta_bombas += 1
            
#--------------------------------------------------------------------------------------------------------------------------------------------------#

def preencherNumeros(qtd_linhas, qtd_colunas, tabuleiro):

    for linha in range(0, qtd_linhas):
        for coluna in range(0, qtd_colunas):

            if tabuleiro[linha][coluna] != '*':
                contaBombas = 0
                
                # 1 - Posição Superior Centro
                if (linha-1>=0 and tabuleiro[linha-1][coluna] == '*'):
                    contaBombas += 1

                # 2 - Posição Superior Direita
                if (linha-1>=0 and coluna+1<qtd_colunas and tabuleiro[linha-1][coluna+1] == '*'):
                    contaBombas += 1

                # 3 - Posição Lateral Direita
                if (coluna+1<qtd_colunas and tabuleiro[linha][coluna+1] == '*'):
                    contaBombas += 1

                # 4 - Posição Inferior Direita
                if (linha+1<qtd_linhas and coluna+1<qtd_colunas and tabuleiro[linha+1][coluna+1] == '*'):
                    contaBombas += 1

                # 5 - Posição Inferior Centro
                if (linha+1<qtd_linhas and tabuleiro[linha+1][coluna] == '*'):
                    contaBombas += 1

                # 6 - Posição Inferior Esquerda
                if (linha+1<qtd_linhas and coluna-1>=0 and tabuleiro[linha+1][coluna-1] == '*'):
                    contaBombas += 1

                # 7 - Posição Lateral Esquerda
                if (coluna-1>=0 and tabuleiro[linha][coluna-1] == '*'):
                    contaBombas += 1

                # 8 - Posição Superior Esquerda
                if (linha-1>=0 and coluna-1>=0 and tabuleiro[linha-1][coluna-1] == '*'):
                    contaBombas += 1

                tabuleiro[linha][coluna] = contaBombas

        print()

#--------------------------------------------------------------------------------------------------------------------------------------------------#        

def CriaTabBD(tabuleiro):
    from openpyxl import Workbook
    arquivo_excel = Workbook()
    Tabuleiro1 = arquivo_excel.active 
    Tabuleiro1.title = 'CAMPO'

    for linha in range(0, len(tabuleiro)):
        Tabuleiro1.append(tabuleiro[linha])

    arquivo_excel.save('Campo_Minado.xlsx')
    
    return 'Campo_Minado.xlsx' 

#--------------------------------------------------------------------------------------------------------------------------------------------------#

def imprimirTabuleiro(qtd_linhas, qtd_colunas):
    from openpyxl import load_workbook
    campo = load_workbook('Campo_Minado.xlsx', read_only=False)
    abas = campo.sheetnames
    tabuleiro = campo[abas[0]]

    TabVisaodeJogo = []

    for lin in range(1, qtd_linhas+1):
        linha = []
        for col in range(1, qtd_colunas+1):
            linha.append('-')
        TabVisaodeJogo.append(linha)

    campo.save('Campo_Minado.xlsx')
    return TabVisaodeJogo

#--------------------------------------------------------------------------------------------------------------------------------------------------#

def jogar(qtd_linhas, qtd_colunas, visaodejogo):
    from openpyxl import load_workbook
    campo = load_workbook('Campo_Minado.xlsx', read_only=False)
    abas = campo.sheetnames
    tabuleiro = campo[abas[0]]
    perdeu = False
    jogadas = []
    ContJogadas = False
    verif = 0
    
    while True:
        colunas = []
        for i in range(1, qtd_colunas+1):
            colunas.append([i])
    
        print(' # ', colunas)
    
        impressor = []
        linhas = [] 
        for lin in range(1, qtd_linhas+1):
            linhas.append(lin)

        for i in  visaodejogo:
            impressor.append(i)
        for elem in linhas:
            print([elem] + [i])
            
        for cord in jogadas:
            l =[]

        EscolheLinha = int(input('Digite a linha da coordenada que deseja: '))
        EscolheColuna = int(input('Digite a coluna da coordenada que deseja: '))  
        
        if(verif == 1):
            if(cord == tabuleiro.cell(row=EscolheLinha, column=EscolheColuna)):
                print('Coordenada já inserida, tente outra!')
                continue
            else:
                if(verif == 1):
                    jogadas.append(tabuleiro.cell(row=EscolheLinha, column=EscolheColuna))
                    for i in range(1, qtd_linhas+1):
                        #lin = 0
                        for j in range(1, qtd_colunas+1):
                            passou = 0
                            for k in jogadas:
                                if(k == tabuleiro.cell(row=i, column=j) and tabuleiro.cell(row=EscolheLinha, column=EscolheColuna).value == '*'):
                                    print(tabuleiro.cell(row=i, column=j).value, end='  ')
                                    passou = 1
                                    perdeu = True
                                    break
                                elif(k == tabuleiro.cell(row=i, column=j) and tabuleiro.cell(row=EscolheLinha, column=EscolheColuna).value != '*'):
                                    print(tabuleiro.cell(row=i, column=j).value, end='   | ')
                                    passou = 1
                            if passou == 0:
                                print ('[ - ]', end=' | ')
                        print()
                    if perdeu == True:
                        print('Game Over!')
                        break
                    else:
                        print()
                    continue
        while True:
            if(ContJogadas == False):
                jogadas.append(tabuleiro.cell(row=EscolheLinha, column=EscolheColuna))
                verif = 1
                break
        ContJogadas = True

        for i in range(1, qtd_linhas+1):
            #lin = 0
            for j in range(1, qtd_colunas+1):
                passou = 0
                for k in jogadas:
                    if(k == tabuleiro.cell(row=i, column=j) and tabuleiro.cell(row=EscolheLinha, column=EscolheColuna).value == '*'):
                        print(tabuleiro.cell(row=i, column=j).value, end='  ')
                        passou = 1
                        perdeu = True
                        break
                    elif(k == tabuleiro.cell(row=i, column=j) and tabuleiro.cell(row=EscolheLinha, column=EscolheColuna).value != '*'):
                        print(tabuleiro.cell(row=i, column=j).value, end='   | ')
                        passou = 1
                if passou == 0:
                    print ('[ - ]', end=' | ')
            print()
        campo.save('Campo_Minado.xlsx')   
        if perdeu == True:
            print('Game Over!')
            break
        else:
            print()
        continue
#--------------------------------------------------------------------------------------------------------------------------------------------------#

         #*********************************************   # ÍNICIO DO PROGRAMA #  *********************************************#

while(True):

    # Passo 01 - Definir tamanho do tabuleiro
    qtd_linhas, qtd_colunas = definirTamanhoTabuleiro()

    # Passo 02 - Criar e configurar o tabuleiro na memória RAM
    
    # Passo 02.1 - Definir nível e calculando a qtd de bombas
    qtd_bombas = definirQtdBombas(qtd_linhas, qtd_colunas)

    # Passo 02.2 - Criar tabuleiro
    tabuleiro = criarTabuleiro(qtd_linhas, qtd_colunas)

    # Passo 02.3 - Lançar bombas no tabuleiro
    lancarbombas = lancarBombas(qtd_linhas, qtd_colunas, qtd_bombas, tabuleiro)

    # Passo 02.4 - Preencher números no tabuleiro
    Numeros = preencherNumeros(qtd_linhas, qtd_colunas, tabuleiro)
    
    # Passo 03 - Colocar no Excel
    TabExcel = CriaTabBD(tabuleiro)
    
    # Passo 04 - Jogar
    visaodejogo = imprimirTabuleiro(qtd_linhas, qtd_colunas)
    
    # Passo 04.2 - Solicitar ao usuário uma posição
    JogarCoordenada = jogar(qtd_linhas, qtd_colunas, visaodejogo)
    
    # Passo X - Deseja sair

    resp = input('Deseja jogar novamente? (S ou N)')

    if resp in 'nN':
        break
  

          #*********************************************   # Fim DO PROGRAMA #  *********************************************#
#--------------------------------------------------------------------------------------------------------------------------------------------------#
